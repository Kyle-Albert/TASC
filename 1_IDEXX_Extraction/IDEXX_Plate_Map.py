# -*- coding: utf-8 -*-

import math
import openpyxl
import os
import sys
from datetime import datetime
import shutil
import time

"""
Get directories and lists together for the files in the input, output, and 'root'
folders
"""
currentdir = os.getcwd()

inputdir = os.getcwd() + "\\Input"
inputlist = os.listdir(inputdir)

outputdir = os.getcwd() + "\\Output"
outputlist = os.listdir(outputdir)

archivedir = os.getcwd() + "\\Archive"





"""
Print messages to the user to indicate how the script functions
"""
print('IDEXX_Plate_Map.py v1.0')

print('')

print('To use this program, open the Notepad and scan the barcodes into the notepad in the order they will go on the extraction plate')
print('A1 = 1, B1 = 2, ... A2 = 9, B2 = 10, ... D12 = 92')
print('Save the name of the text file from the notepad as the extraction plate barcode')

for i in range(2):
    print('')

print('This program converts a text file (.txt) from scanning the barcodes')
print('Into an excel (.xlsx) visual plate layout is created')

for i in range(2):
    print('')


print('Ensure the proper text file was put into the input folder within the 1_IDEXX_Extraction folder')
print(inputdir)

for i in range(5):
    print('')
    
    
    
    
    
"""
Select the file you want to use
"""
numberdir = []
for i in range(len(inputlist)):
    print(str(i) + " : "  + inputlist[i])
    numberdir.append(str(i))
filenumber = input("Please input the number associated with the file:")

while filenumber not in numberdir:
    for i in range(6):
        print('')

    for i in range(len(inputlist)):
        print(str(i) + " : "  + inputlist[i])
    filenumber = input("Please input the number associated with the file:")
    
filenumber = int(filenumber)

textfile = inputlist[filenumber]
extractionbarcode = textfile[0 : len(textfile) - 4]

barcodetxtfile = open(inputdir + "//" + textfile, "r")#reads the barcode scans file from the desired date
barcodes = [(line.strip()).split() for line in barcodetxtfile]#makes each column of the file (each barcode) an element in a list
barcodetxtfile.close()#closes the file

barcode_len = len(barcodes)#determines how many barcodes are in the list






"""
Duplicate barcode check
"""
duplicatebarcodes = []
for i in range(barcode_len):
    jindex = []
    for j in range(barcode_len):
        if barcodes[i] == barcodes[j] and i < j:
            jindex.append(str(j + 1))
        if j == barcode_len - 1 and jindex != [] and (barcodes[i] not in duplicatebarcodes):
            duplicatebarcodes.append(barcodes[i])
            print(str(str(barcodes[i])[2:len(str(barcodes[i]))-2]) + " has duplicates at positions " + str(i + 1) + ", " + ", ".join(jindex))

if duplicatebarcodes != []:
    print("Duplicate barcodes detected")
    exitscript = input("Would you like to continue? (y/n):")
    while exitscript != "y" or exitscript != "n":
        if exitscript == "y":
            break
        if exitscript == "n":
            sys.exit()
        exitscript = input("Would you like to continue? (y/n):")
else:
    print("No duplicate barcodes")





"""
Delete Blank spaces entered in the textfile
Recreate the barcode_len variable with updated length
"""
deletenum = 0
for i in range(barcode_len):
    if len(barcodes[i]) == 0:
        barcodes[i] = ''
        deletenum += 1
        
for i in range(deletenum):
    barcodes.remove('')

barcode_len = len(barcodes)





"""
Class implementation for barcodes
"""
class Barcodes:
    def __init__(self, wellid, sample_barcode, extractionplatebarcode):
        self.wellid = wellid
        self.sample_barcode = sample_barcode
        self.extractionplatebarcode = extractionplatebarcode
        pass

rows = ["A", "B", "C", "D", "E", "F", "G", "H"]
columns = ["1","2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12"]
dictionary = []
for column in columns:
    for row in rows:
        dictionary.append(row + column)

for i in range(barcode_len):
    barcodes[i][0] = Barcodes(wellid = dictionary[i], sample_barcode = barcodes[i][0], extractionplatebarcode = extractionbarcode)






"""
Excel
"""
wb = openpyxl.load_workbook('IDEXX Template - TASC Extraction Plate.xlsx')
ws = wb['Extraction Plate']
ws2 = wb['Scans']
ws.cell(1 , 4, extractionbarcode)
ws.cell(1 , 9, datetime.now().strftime("%m/%d/%Y"))



for i in range(barcode_len):
    column = 2 + math.floor(i / 8)
    row = i+3 -(8 * math.floor(i / 8))
    ws.cell(row, column, str(barcodes[i][0].sample_barcode))


for i in range(barcode_len):
    ws2.cell(i+1, 1, str(barcodes[i][0].sample_barcode))
wb.save(extractionbarcode + ' - TASC Extraction Plate.xlsx')







shutil.move(extractionbarcode + ' - TASC Extraction Plate.xlsx', outputdir)
exceldir = outputdir + '\\' + extractionbarcode + ' - TASC Extraction Plate.xlsx'
os.startfile(exceldir, 'print')
time.wait(5)
os.close(exceldir)
os.remove(exceldir)

