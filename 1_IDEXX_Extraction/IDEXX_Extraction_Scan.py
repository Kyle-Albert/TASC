# -*- coding: utf-8 -*-

import math
import openpyxl
import os
import sys
from datetime import datetime
import csv
import shutil
import time



"""
Get directories and list for input, output, and root
"""
currentdir = os.getcwd()

inputdir = os.getcwd() + "\\Input"
inputlist = os.listdir(inputdir)

outputdir = os.getcwd() + "\\Output"
outputlist = os.listdir(outputdir)

archivedir = os.getcwd() + "\\Archive"
archivelist = os.listdir(archivedir)
print('IDEXX_Extraction_Scan.py v1.3')
print('')


"""
Detect duplicate files in input/output with archive folder, and stop if any are detected
"""
duplicatefilecount = 0
for i in range(len(outputlist)):
    for j in range(len(archivelist)):
        if outputlist[i] == archivelist[j]:
            duplicatefilecount += 1
            print('"' + outputlist[i] + '" exists in both the archive and OUTPUT folders. The program will not proceed until "' + outputlist[i] + '" is removed from the archive folder' )
            for k in range(2):
                print("")

for i in range(len(inputlist)):
    for j in range(len(archivelist)):
        if inputlist[i] == archivelist[j]:
            duplicatefilecount += 1
            print('"' + inputlist[i] + '" exists in both the archive and INPUT folders. The program will not proceed until "' + inputlist[i] + '" is removed from the archive folder' )
            for k in range(2):
                print("")

if duplicatefilecount > 0:
    print("The program will close in 20 seconds. Please correct the duplicate files and execute the program again.")
    time.sleep(20)
    sys.exit()




"""
Print messages to the user to indicate how the script functions
"""
print('To use this program, open the Notepad and scan the barcodes into the notepad in the order they will go on the extraction plate')
print('A1 = 1, B1 = 2, ... A2 = 9, B2 = 10, ... D12 = 92')
print('Save the name of the text file from the notepad as the extraction plate barcode')

for i in range(2):
    print('')

print('This program converts a text file (.txt) from scanning the barcodes')
print('to a CSV file (.csv) containing the extraction plate layout, to be uploaded onto LIMS')
print('Additonally, an excel (.xlsx) visual plate layout is created')

for i in range(2):
    print('')


print('Ensure the proper text file was put into the input folder within the 1_IDEXX_Extraction folder')
print(inputdir)

print('')

print('The extraction plate CSV file and visual plate layout excel file will be in the output folder within the 1_IDEXX_Extraction folder')
print(outputdir)

print('')

print('The text file selected from the input folder will be moved to the archive folder after the program is complete')
print('The contents that were in the output folder from previous executions will be moved to the archive folder')

print('')

print('If the program fails to work, ensure the text file you would like to use is not present in both the input and archive folders')
print('Additonally, ensure there are not files present in the both the output and archive folders with the same name')
print('Do not run the same file twice as it will cause errors. It this occurs, delete the unwanted outputs from the archive and delete the')
print('resulting outputs, then run the proper file.')

for i in range(5):
    print('')




"""
Select the file you want to use
"""
numberdir = []
for i in range(len(inputlist)):
    print(str(i) + " : "  + inputlist[i])
    numberdir.append(str(i))
filenumber = input("Please input the number associated with the file:")

while filenumber not in numberdir:
    for i in range(6):
        print('')

    for i in range(len(inputlist)):
        print(str(i) + " : "  + inputlist[i])
    filenumber = input("Please input the number associated with the file:")
    
filenumber = int(filenumber)

textfile = inputlist[filenumber]
extractionbarcode = textfile[0 : len(textfile) - 4]



barcodetxtfile = open(inputdir + "//" + textfile, "r")#reads the barcode scans file from the desired date
barcodes = [(line.strip()).split() for line in barcodetxtfile]#makes each column of the file (each barcode) an element in a list
barcodetxtfile.close()#closes the file

barcode_len = len(barcodes)#determines how many barcodes are in the list



"""
Delete Blank spaces entered in the textfile
Recreate the barcode_len variable with updated length
"""
deletenum = 0
for i in range(barcode_len):
    if len(barcodes[i]) == 0:
        barcodes[i] = ''
        deletenum += 1
        
for i in range(deletenum):
    barcodes.remove('')

barcode_len = len(barcodes)



"""
Duplicate barcode check
"""
duplicatebarcodes = []
for i in range(barcode_len):
    jindex = []
    for j in range(barcode_len):
        if barcodes[i] == barcodes[j] and i < j:
            jindex.append(str(j + 1))
        if j == barcode_len - 1 and jindex != [] and (barcodes[i] not in duplicatebarcodes):
            duplicatebarcodes.append(barcodes[i])
            print(str(str(barcodes[i])[2:len(str(barcodes[i]))-2]) + " has duplicates at positions " + str(i + 1) + ", " + ", ".join(jindex))

if duplicatebarcodes != []:
    print("Duplicate barcodes detected")
    exitscript = input("Would you like to continue? (y/n):")
    while exitscript != "y" or exitscript != "n":
        if exitscript == "y":
            break
        if exitscript == "n":
            sys.exit()
        exitscript = input("Would you like to continue? (y/n):")
else:
    print("No duplicate barcodes")



"""
Class implementation for barcodes
"""
class Barcodes:
    def __init__(self, wellid, sample_barcode, extractionplatebarcode):
        self.wellid = wellid
        self.sample_barcode = sample_barcode
        self.extractionplatebarcode = extractionplatebarcode
        pass

rows = ["A", "B", "C", "D", "E", "F", "G", "H"]
columns = ["1","2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12"]
dictionary = []
for column in columns:
    for row in rows:
        dictionary.append(row + column)

for i in range(barcode_len):
    barcodes[i][0] = Barcodes(wellid = dictionary[i], sample_barcode = barcodes[i][0], extractionplatebarcode = extractionbarcode)
    




"""
PCR Template
This is for if a pcr file would need to be generated from this script, from a
text file, instead of from the csv.


PCRtxtfile = open("IDEXX Template.txt", "r")
list_of_lines = PCRtxtfile.readlines()#reads the lines from PCR template file and replaces all the "replace" spots with the correct barcodes

for i in range(barcode_len): 
    for j in range(len(list_of_lines)):
        number = list_of_lines[j][1:3]
        if number == "10" or number == "11" or number == "12":
            PCRwellid = list_of_lines[j][:3]
        else:
            PCRwellid = list_of_lines[j][:2]
        
        if barcodes[i][0].wellid == PCRwellid:
            list_of_lines[j] = list_of_lines[j].replace("replace", barcodes[i][0].sample_barcode)

for i in range(len(list_of_lines)):
    if "replace" in list_of_lines[i]:
        list_of_lines[i] = ''

PCRtxtfile = open("IDEXX Template Edited.txt", "w")#This is the new file that we are overwriting onto
PCRtxtfile.writelines(list_of_lines)#since we are changing 3 lines at a time throughout the entire list, we want to save those each time a loop occurs
PCRtxtfile.close()#closing the file is what actually saves the changes
"""




"""
Excel
"""
wb = openpyxl.load_workbook('IDEXX Template - TASC Extraction Plate.xlsx')
ws = wb['Extraction Plate']
ws2 = wb['Scans']
ws.cell(1 , 4, extractionbarcode)
ws.cell(1 , 9, datetime.now().strftime("%m/%d/%Y"))


for i in range(barcode_len):
    column = 2 + math.floor(i / 8)
    row = i+3 -(8 * math.floor(i / 8))
    ws.cell(row, column, str(barcodes[i][0].sample_barcode))


for i in range(barcode_len):
    ws2.cell(i+1, 1, str(barcodes[i][0].sample_barcode))
wb.save(extractionbarcode + ' - TASC Extraction Plate.xlsx')




"""
LIMS CSV Export
"""
with open(extractionbarcode + ' - LIMS Extraction CSV.csv', 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['Well ID', 'Sample Barcode', 'Extraction Plate Barcode'])
    for i in range(barcode_len):
        b = barcodes[i][0]
        writer.writerow([b.wellid, b.sample_barcode, b.extractionplatebarcode])
f.close()


print('The CSV of the extraction plate, ' + extractionbarcode + ' - LIMS Extraction CSV.csv, has been successfully created.')
for i in range(2):
    print('')


"""
Move files to output folders and archives, print extraction plate layout
"""

for i in range(len(outputlist)):
    shutil.move(outputdir + "//" + outputlist[i], archivedir)

shutil.move(extractionbarcode + ' - LIMS Extraction CSV.csv', outputdir)
shutil.move(extractionbarcode + ' - TASC Extraction Plate.xlsx', outputdir)
shutil.move(inputdir + "//" + textfile, archivedir)

exceldir = outputdir + '\\' + extractionbarcode + ' - TASC Extraction Plate.xlsx'
os.startfile(exceldir, 'print')
os.startfile(exceldir)

print('The CSV of the extraction plate is in the output folder and is ready for upload to LIMS')

time.sleep(10)
